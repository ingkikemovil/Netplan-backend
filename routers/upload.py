from fastapi import APIRouter, UploadFile, File, HTTPException
import openpyxl
import io

router = APIRouter()


@router.post("/excel")
async def parse_excel(file: UploadFile = File(...)):
    """
    Lee un archivo Excel con dos hojas:
      - Hoja 'nodos': columnas nombre, tipo, latitud, longitud, direccion
      - Hoja 'conexiones': columnas origen, destino, costo, tipo
    Devuelve los datos parseados listos para insertar.
    """
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos .xlsx")

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

    nodes = []
    connections = []

    if "nodos" in wb.sheetnames:
        ws = wb["nodos"]
        headers = [str(cell.value).strip().lower() for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            row_dict = dict(zip(headers, row))
            nodes.append({
                "name": str(row_dict.get("nombre", row_dict.get("name", ""))),
                "node_type": str(row_dict.get("tipo", row_dict.get("type", "city"))).lower(),
                "latitude": float(row_dict.get("latitud", row_dict.get("latitude", 0))),
                "longitude": float(row_dict.get("longitud", row_dict.get("longitude", 0))),
                "address": str(row_dict.get("direccion", row_dict.get("address", ""))) or None,
            })

    if "conexiones" in wb.sheetnames:
        ws = wb["conexiones"]
        headers = [str(cell.value).strip().lower() for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            row_dict = dict(zip(headers, row))
            connections.append({
                "source_name": str(row_dict.get("origen", row_dict.get("source", ""))),
                "target_name": str(row_dict.get("destino", row_dict.get("target", ""))),
                "cost": float(row_dict.get("costo", row_dict.get("cost", 0))),
                "conn_type": str(row_dict.get("tipo", row_dict.get("type", "normal"))).lower(),
            })

    return {"nodes": nodes, "connections": connections}


@router.get("/template")
def download_template_info():
    return {
        "instrucciones": "Descarga la plantilla Excel desde el frontend. Tiene dos hojas:",
        "hoja_nodos": ["nombre", "tipo", "latitud", "longitud", "direccion"],
        "hoja_conexiones": ["origen", "destino", "costo", "tipo"],
        "tipos_nodo": ["city", "tower", "datacenter", "other"],
        "tipos_conexion": ["normal", "mandatory", "forbidden"],
    }
